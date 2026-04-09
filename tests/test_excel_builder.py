"""Tests for excel_builder module."""

import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from tear_sheet.charts import create_price_chart, create_revenue_chart
from tear_sheet.excel_builder import build_tear_sheet


def _make_sample_data():
    """Create a complete mock data dict matching scraper output."""
    return {
        "info": {
            "company_name": "Test Corp",
            "ticker": "TEST",
            "sector": "Industrials",
            "industry": "Machinery",
            "description": "A test company that makes things.",
            "city": "Peoria",
            "state": "IL",
            "employees": 100000,
            "forward_pe": 15.5,
            "trailing_pe": 18.2,
            "market_cap": 150e9,
            "fifty_two_week_high": 350.0,
            "fifty_two_week_low": 250.0,
            "current_ratio": 1.4,
            "quick_ratio": 0.9,
            "debt_to_equity": 1.8,
            "beta": 1.1,
            "dividend_yield": 0.017,
            "payout_ratio": 0.25,
            "trailing_eps": 18.5,
            "forward_eps": 20.0,
            "recommendation_key": "buy",
            "current_price": 310.0,
        },
        "income_stmt": _make_income_stmt(),
        "balance_sheet": None,
        "history": _make_history(),
        "recommendations": None,
        "analyst_targets": {"low": 280, "mean": 360, "median": 355, "high": 400, "current": 310},
        "calendar": {"Earnings Date": ["2026-04-25"]},
        "ticker": "TEST",
        "fetch_date": "2026-04-09",
    }


def _make_income_stmt():
    dates = pd.to_datetime(["2021-12-31", "2022-12-31", "2023-12-31", "2024-12-31"])
    data = {
        "Total Revenue": [50e9, 55e9, 60e9, 63e9],
        "Net Income": [5e9, 6e9, 5.5e9, 7e9],
    }
    return pd.DataFrame(data, index=dates).T


def _make_history():
    dates = pd.bdate_range("2020-01-01", periods=1200)
    np.random.seed(42)
    prices = 250 + np.cumsum(np.random.randn(1200) * 0.5)
    return pd.DataFrame({"Close": prices}, index=dates)


def _make_financials():
    return {
        "growth": {
            "revenue_yoy": {"2021→2022": 0.1, "2022→2023": 0.0909, "2023→2024": 0.05},
            "revenue_indexed": pd.Series([1.0, 1.1, 1.2, 1.26]),
            "net_income_yoy": {"2021→2022": 0.2, "2022→2023": -0.083, "2023→2024": 0.273},
        },
        "returns": {"1Y": 0.15, "3Y": 0.35, "5Y": 0.60},
        "ratios": {"current_ratio": 1.4, "quick_ratio": 0.9, "debt_to_equity": 1.8},
        "analyst_upside": 0.161,
    }


def _make_charts(data):
    charts = {}
    if data["income_stmt"] is not None:
        charts["revenue_chart"] = create_revenue_chart(data["income_stmt"], "TEST")
    if data["history"] is not None:
        charts["price_chart"] = create_price_chart(data["history"], "TEST")
    return charts


class TestBuildTearSheet:
    def test_produces_valid_xlsx(self):
        data = _make_sample_data()
        financials = _make_financials()
        charts = _make_charts(data)

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "TEST_tear_sheet.xlsx"
            result = build_tear_sheet(data, financials, charts, path)

            assert result.exists()
            assert result.suffix == ".xlsx"

            # Verify it's a valid workbook
            wb = load_workbook(result)
            ws = wb.active
            assert ws.title == "Tear Sheet"

            # Check title
            assert ws.cell(row=1, column=1).value == "Test Corp"
            assert ws.cell(row=1, column=18).value == "TEST"

            # Check section headers exist (row numbers from SECTION_ROWS in config)
            assert ws.cell(row=3, column=1).value == "COMPANY OVERVIEW"
            assert ws.cell(row=12, column=1).value == "FINANCIAL DATA"
            assert ws.cell(row=28, column=1).value == "STOCK PERFORMANCE"
            assert ws.cell(row=45, column=1).value == "VALUATION"
            assert ws.cell(row=54, column=1).value == "SOURCES"

    def test_handles_missing_data(self):
        """Test with minimal data — no charts, no analyst targets."""
        data = _make_sample_data()
        data["income_stmt"] = None
        data["history"] = None
        data["analyst_targets"] = None
        data["calendar"] = None

        financials = {
            "growth": {"revenue_yoy": {}, "revenue_indexed": None, "net_income_yoy": {}},
            "returns": {"1Y": None, "3Y": None, "5Y": None},
            "ratios": {"current_ratio": None, "quick_ratio": None, "debt_to_equity": None},
            "analyst_upside": None,
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "TEST_tear_sheet.xlsx"
            result = build_tear_sheet(data, financials, {}, path)
            assert result.exists()

            wb = load_workbook(result)
            ws = wb.active
            assert ws.cell(row=1, column=1).value == "Test Corp"
