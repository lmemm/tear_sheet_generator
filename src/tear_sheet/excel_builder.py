"""Excel tear sheet assembly and formatting."""

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.properties import PageSetupProperties
from tear_sheet.config import (
    COLUMN_WIDTHS,
    DESCRIPTION_ROW_HEIGHT,
    FONT_NAME,
    HEADER_BG_COLOR,
    HEADER_FONT_COLOR,
    HEADER_FONT_SIZE,
    LABEL_FONT_SIZE,
    LAST_COL_NUM,
    PAGE_MARGINS,
    SECTION_ROWS,
    TITLE_FONT_SIZE,
    VALUE_FONT_SIZE,
)
from tear_sheet.utils import format_currency, format_market_cap, format_percentage


def build_tear_sheet(
    data: dict, financials: dict, charts: dict, output_path: Path
) -> Path:
    """Assemble all data and charts into a formatted Excel tear sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tear Sheet"

    info = data.get("info", {})
    growth = financials.get("growth", {})
    returns = financials.get("returns", {})
    ratios = financials.get("ratios", {})
    analyst_upside = financials.get("analyst_upside")

    # -- Column widths ----------------------------------------------------
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # =====================================================================
    # Section 1: Company Overview
    # =====================================================================
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    title_cell = ws.cell(row=1, column=1, value=_val(info.get("company_name")))
    title_cell.font = Font(name=FONT_NAME, size=TITLE_FONT_SIZE, bold=True)
    title_cell.alignment = Alignment(horizontal="center")
    ticker_cell = ws.cell(row=1, column=18, value=_val(info.get("ticker")))
    ticker_cell.font = Font(name=FONT_NAME, size=TITLE_FONT_SIZE, bold=True)
    ticker_cell.alignment = Alignment(horizontal="right")

    # Section header
    row = SECTION_ROWS["company_overview"]
    _write_section_header(ws, row, "COMPANY OVERVIEW")

    # Row 4: Industry | Sector
    _write_label_value(ws, row + 1, 2, "Industry:", 3, info.get("industry"))
    _write_label_value(ws, row + 1, 10, "Sector:", 11, info.get("sector"))

    # Row 5: Location | Employees
    city = info.get("city") or ""
    state = info.get("state") or ""
    location = f"{city}, {state}".strip(", ") if (city or state) else None
    employees = info.get("employees")
    emp_display = f"{employees:,}" if employees else "N/A"
    _write_label_value(ws, row + 2, 2, "Location:", 3, location)
    _write_label_value(ws, row + 2, 10, "Employees:", 11, emp_display)

    # Rows 6-11: Description (expanded to avoid cutoff)
    _write_label_value(ws, row + 3, 2, "Description:", 3, None)
    desc = info.get("description") or "N/A"
    ws.merge_cells(start_row=row + 3, start_column=3, end_row=row + 8, end_column=LAST_COL_NUM)
    desc_cell = ws.cell(row=row + 3, column=3, value=desc)
    desc_cell.font = Font(name=FONT_NAME, size=VALUE_FONT_SIZE)
    desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
    # Set row heights so description text is fully visible
    for desc_r in range(row + 3, row + 9):
        ws.row_dimensions[desc_r].height = DESCRIPTION_ROW_HEIGHT

    # =====================================================================
    # Section 2: Financial Data
    # =====================================================================
    row = SECTION_ROWS["financial_data"]
    _write_section_header(ws, row, "FINANCIAL DATA")

    # Revenue chart (I14 to Q26)
    if "revenue_chart" in charts:
        _embed_image(ws, charts["revenue_chart"], "I14", width=680, height=245)

    # Revenue Growth YoY (pushed down to clear chart)
    yoy_row = row + 10
    _write_label_value(ws, yoy_row, 2, "Revenue Growth (YoY):", 3, None)
    rev_yoy = growth.get("revenue_yoy", {})
    col = 3
    for period, pct in rev_yoy.items():
        ws.cell(row=yoy_row, column=col, value=f"{period}: {format_percentage(pct)}")
        ws.cell(row=yoy_row, column=col).font = Font(name=FONT_NAME, size=VALUE_FONT_SIZE)
        col += 2

    # Net Income Growth YoY
    ni_row = yoy_row + 1
    _write_label_value(ws, ni_row, 2, "Net Income Growth (YoY):", 3, None)
    ni_yoy = growth.get("net_income_yoy", {})
    col = 3
    for period, pct in ni_yoy.items():
        ws.cell(row=ni_row, column=col, value=f"{period}: {format_percentage(pct)}")
        ws.cell(row=ni_row, column=col).font = Font(name=FONT_NAME, size=VALUE_FONT_SIZE)
        col += 2

    # Ratios
    ratio_row = ni_row + 1
    cr = ratios.get("current_ratio")
    qr = ratios.get("quick_ratio")
    _write_label_value(ws, ratio_row, 2, "Current Ratio:", 3, f"{cr:.2f}" if cr else "N/A")
    _write_label_value(ws, ratio_row, 6, "Quick Ratio:", 7, f"{qr:.2f}" if qr else "N/A")

    dte = ratios.get("debt_to_equity")
    _write_label_value(
        ws, ratio_row + 1, 2, "Debt-to-Equity:", 3, f"{dte:.2f}" if dte else "N/A"
    )

    # =====================================================================
    # Section 3: Stock Performance
    # =====================================================================
    row = SECTION_ROWS["stock_performance"]
    _write_section_header(ws, row, "STOCK PERFORMANCE")

    # Price chart (I30 to Q43)
    if "price_chart" in charts:
        _embed_image(ws, charts["price_chart"], "I30", width=680, height=260)

    # Returns (stacked with 1-row spacing so they don't stretch under the chart)
    ret_row = row + 2
    _write_label_value(ws, ret_row, 2, "1-Year Return:", 3, format_percentage(returns.get("1Y")))
    _write_label_value(ws, ret_row + 2, 2, "3-Year Return:", 3, format_percentage(returns.get("3Y")))
    _write_label_value(ws, ret_row + 4, 2, "5-Year Return:", 3, format_percentage(returns.get("5Y")))

    # Beta
    beta = info.get("beta")
    _write_label_value(ws, ret_row + 6, 2, "Beta:", 3, f"{beta:.2f}" if beta else "N/A")

    # Dividend / Payout
    _write_label_value(
        ws, ret_row + 7, 2, "Dividend Yield:", 3, format_percentage(info.get("dividend_yield"))
    )
    _write_label_value(
        ws, ret_row + 7, 5, "Payout Ratio:", 6, format_percentage(info.get("payout_ratio"))
    )

    # =====================================================================
    # Section 4: Valuation
    # =====================================================================
    row = SECTION_ROWS["valuation"]
    _write_section_header(ws, row, "VALUATION")

    # P/E
    fpe = info.get("forward_pe")
    tpe = info.get("trailing_pe")
    _write_label_value(ws, row + 1, 2, "Forward P/E:", 3, f"{fpe:.2f}" if fpe else "N/A")
    _write_label_value(ws, row + 1, 6, "Trailing P/E:", 7, f"{tpe:.2f}" if tpe else "N/A")

    # 52-week
    _write_label_value(
        ws, row + 2, 2, "52-Week High:", 3, format_currency(info.get("fifty_two_week_high"))
    )
    _write_label_value(
        ws, row + 2, 6, "52-Week Low:", 7, format_currency(info.get("fifty_two_week_low"))
    )

    # Market cap
    _write_label_value(ws, row + 3, 2, "Market Cap:", 3, format_market_cap(info.get("market_cap")))

    # EPS
    _write_label_value(
        ws, row + 4, 2, "Trailing EPS:", 3, format_currency(info.get("trailing_eps"))
    )
    _write_label_value(
        ws, row + 4, 6, "Forward EPS:", 7, format_currency(info.get("forward_eps"))
    )

    # Analyst recommendation
    rec = info.get("recommendation_key") or "N/A"
    _write_label_value(ws, row + 5, 2, "Analyst Recommendation:", 3, rec.title() if rec != "N/A" else rec)

    targets = data.get("analyst_targets")
    mean_target = targets.get("mean") if targets else None
    _write_label_value(
        ws, row + 5, 8, "Mean Price Target:", 9, format_currency(mean_target)
    )

    # Upside/Downside
    upside_str = format_percentage(analyst_upside)
    upside_cell = _write_label_value(ws, row + 6, 2, "Upside/Downside:", 3, upside_str)
    if analyst_upside is not None:
        color = "228B22" if analyst_upside >= 0 else "CC0000"
        upside_cell.font = Font(name=FONT_NAME, size=VALUE_FONT_SIZE, color=color, bold=True)

    # Next earnings date
    calendar = data.get("calendar")
    earnings_date = None
    if isinstance(calendar, dict):
        ed = calendar.get("Earnings Date")
        if ed:
            earnings_date = str(ed[0]) if isinstance(ed, list) else str(ed)
    _write_label_value(ws, row + 7, 2, "Next Earnings Date:", 3, earnings_date or "N/A")

    # =====================================================================
    # Section 5: Sources / Citations
    # =====================================================================
    row = SECTION_ROWS["sources"]
    _write_section_header(ws, row, "SOURCES")

    ticker_str = data.get("ticker", "")
    sources = [
        ("Yahoo Finance — Company Profile", f"https://finance.yahoo.com/quote/{ticker_str}/"),
        ("Yahoo Finance — Financials", f"https://finance.yahoo.com/quote/{ticker_str}/financials/"),
        ("Yahoo Finance — Analysis", f"https://finance.yahoo.com/quote/{ticker_str}/analysis/"),
        ("SEC EDGAR — Company Filings", f"https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&company={ticker_str}&type=10-K&dateb=&owner=include&count=10"),
    ]
    for i, (label, url) in enumerate(sources):
        src_row = row + 1 + i
        src_cell = ws.cell(row=src_row, column=2, value=label)
        src_cell.font = Font(name=FONT_NAME, size=9, color="4472C4", underline="single")
        src_cell.hyperlink = url
        src_cell.alignment = Alignment(horizontal="left")
        # Merge C-E for the URL so it doesn't stretch columns
        ws.merge_cells(start_row=src_row, start_column=3, end_row=src_row, end_column=5)
        url_cell = ws.cell(row=src_row, column=3, value=url)
        url_cell.font = Font(name=FONT_NAME, size=8, color="808080")
        url_cell.alignment = Alignment(horizontal="left")

    # Fetch date note
    fetch_date = data.get("fetch_date", "N/A")
    note_cell = ws.cell(row=row + len(sources) + 1, column=2, value=f"Data retrieved: {fetch_date}")
    note_cell.font = Font(name=FONT_NAME, size=8, color="808080", italic=True)

    # -- Page setup -------------------------------------------------------
    _apply_page_setup(ws)

    # -- Save -------------------------------------------------------------
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


# =========================================================================
# Helper functions
# =========================================================================

def _write_section_header(ws, row: int, title: str):
    """Write a colored section header bar across the full width."""
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=LAST_COL_NUM
    )
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(
        name=FONT_NAME, size=HEADER_FONT_SIZE, bold=True, color=HEADER_FONT_COLOR
    )
    cell.fill = PatternFill(start_color=HEADER_BG_COLOR, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    # Apply fill to all merged cells in the row
    for col in range(2, LAST_COL_NUM + 1):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill(start_color=HEADER_BG_COLOR, fill_type="solid")


def _write_label_value(ws, row: int, label_col: int, label: str, value_col: int, value):
    """Write a label cell and adjacent value cell. Returns the value cell."""
    label_cell = ws.cell(row=row, column=label_col, value=label)
    label_cell.font = Font(name=FONT_NAME, size=LABEL_FONT_SIZE, bold=True)
    label_cell.alignment = Alignment(horizontal="right")

    display = value if value is not None else "N/A"
    value_cell = ws.cell(row=row, column=value_col, value=display)
    value_cell.font = Font(name=FONT_NAME, size=VALUE_FONT_SIZE)
    value_cell.alignment = Alignment(horizontal="left")
    return value_cell


def _embed_image(ws, buf: BytesIO, anchor: str, width: int, height: int):
    """Embed a BytesIO PNG image into the worksheet at the given anchor."""
    buf.seek(0)
    img = XlImage(buf)
    img.width = width
    img.height = height
    ws.add_image(img, anchor)


def _apply_page_setup(ws):
    """Configure worksheet for print-ready landscape letter output."""
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    ws.page_margins.left = PAGE_MARGINS["left"]
    ws.page_margins.right = PAGE_MARGINS["right"]
    ws.page_margins.top = PAGE_MARGINS["top"]
    ws.page_margins.bottom = PAGE_MARGINS["bottom"]


def _val(v):
    """Return value or 'N/A' if None."""
    return v if v is not None else "N/A"
